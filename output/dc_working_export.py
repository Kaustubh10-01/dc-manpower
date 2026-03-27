"""Generate detailed working Excel for a specific DC.

Full traceability chain:
  Raw Data (source) → Activity sheets (vol × load_factor, time-shifted) →
  Activity sheets (MH = vol × effort) → Summary (sum of all activity MH)

All cells are Excel formulas referencing upstream — change Raw Data and everything cascades.
"""

import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from config import SHIFTS, ALL_DOCK_ACTIVITIES
from core.time_shifter import (
    _build_proportional_offset_map,
    _build_activity_mh_rates,
    VOLUMETRIC_ACTIVITIES,
)


def _fh(h):
    """Format hour number to label."""
    if h == 0:
        return "12 AM"
    elif h < 12:
        return f"{h} AM"
    elif h == 12:
        return "12 PM"
    else:
        return f"{h - 12} PM"


HOUR_LABELS = [_fh(h) for h in range(24)]

HEADER_FONT = Font(bold=True)
BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SPACER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
NUM_VOL = "#,##0"
NUM_MH = "#,##0.00"

# ── Column layout (shared across ALL sheets for alignment) ──
# A=Layout, B=Layout Type, C=Date
# D(4)..AA(27) = hours 0-23 (left block)
# AB(28) = Total (left)
# AC(29) = spacer
# AD(30)..BA(53) = hours 0-23 (right block)
# BB(54) = Total (right)
COL_A, COL_B, COL_C = 1, 2, 3
LEFT_START = 4       # D
LEFT_END = 27        # AA
LEFT_TOTAL = 28      # AB
SPACER = 29          # AC
RIGHT_START = 30     # AD
RIGHT_END = 53       # BA
RIGHT_TOTAL = 54     # BB


def generate_dc_working_excel(
    dc_name, load_df_preship, _act_detail, _shifted,
    peak_days, staffing_df, template, xd_combined_staffing=None,
):
    """Generate detailed working Excel for one DC — ALL dates."""
    pre = load_df_preship[load_df_preship["DC"] == dc_name].copy()
    staff = staffing_df[staffing_df["DC"] == dc_name].copy()

    if pre.empty:
        buf = io.BytesIO()
        wb = Workbook(); ws = wb.active; ws.title = "No Data"
        ws["A1"] = f"No data for {dc_name}"
        wb.save(buf); buf.seek(0)
        return buf.getvalue()

    offset_map = _build_proportional_offset_map(template["activity_prod"])
    lt_rates = _build_activity_mh_rates(
        template["layout_prod_df"], template["activity_names"], template["activity_manhours"]
    )
    vol_rates = {}
    for act in VOLUMETRIC_ACTIVITIES:
        e = template["activity_manhours"].get(act, 0)
        if e > 0:
            vol_rates[act] = float(e)

    load_factors = _build_load_factors(template["layout_prod_df"], template["activity_names"])
    activity_efforts = template["activity_manhours"]

    # XD multipliers for this DC
    xd_sites = template.get("xd_sites", set())
    xd_multipliers = template.get("xd_multipliers", {})
    is_xd = dc_name in xd_sites
    dc_xd_mults = xd_multipliers.get(dc_name, {}) if is_xd else {}

    # Exclude dock activities — those are in the dock pipeline
    dock_exclude = set(ALL_DOCK_ACTIVITIES)

    dc_lts = pre["Layout Type"].unique()
    all_acts = set()
    for lt in dc_lts:
        if lt in lt_rates:
            all_acts.update(a for a in lt_rates[lt].keys() if a not in dock_exclude)
    all_acts.update(a for a in vol_rates.keys() if a not in dock_exclude)
    all_acts = sorted(all_acts)

    # Build ordered row data
    row_data = _build_rows(pre)

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Raw Data sheet (SOURCE OF TRUTH) ──
    _write_raw_data(wb, row_data)

    # ── 2. Per-activity sheets ──
    act_sheets = {}
    for act in all_acts:
        sn = act[:31]
        if sn in act_sheets:
            sn = sn[:28] + "_2"
        act_sheets[sn] = act

        is_vol = act in VOLUMETRIC_ACTIVITIES
        offsets = offset_map.get(act, [(0, 1.0)])
        effort = vol_rates.get(act, 0) if is_vol else activity_efforts.get(act, 0)

        # Apply XD multiplier to effort if applicable
        xd_mult = dc_xd_mults.get(act, 1.0)
        effort_adjusted = effort * xd_mult

        _write_activity_sheet(
            wb, sn, act, row_data, load_factors, effort_adjusted, offsets, is_vol,
            xd_mult=xd_mult if xd_mult != 1.0 else None,
        )

    # ── 3. Summary sheet (sum of all activity MH) ──
    _write_summary(wb, row_data, act_sheets, is_xd=is_xd)

    # ── 4. Staffing sheet ──
    # For XD sites, show combined staffing if available
    xd_staff = None
    if is_xd and xd_combined_staffing is not None and not xd_combined_staffing.empty:
        xd_staff = xd_combined_staffing[xd_combined_staffing["DC"] == dc_name]
    _write_staffing(wb, staff, xd_staff=xd_staff)

    # Reorder: Summary first, then Raw Data, then activities, then Staffing
    desired = ["Summary", "Raw Data"] + sorted(act_sheets.keys()) + ["Staffing"]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════

def _build_load_factors(layout_prod_df, activity_names):
    factors = {}
    for _, row in layout_prod_df.iterrows():
        lt = row["Layout Type"]
        lt_f = {}
        for act in activity_names:
            lf = row.get(act, 0)
            if pd.isna(lf) or lf == 0:
                continue
            lt_f[act] = float(lf)
        factors[lt] = lt_f
    return factors


def _build_rows(pre):
    """Build ordered list of dicts for consistent row indexing across all sheets."""
    grp = pre.groupby(["layout_name", "Layout Type", "Date of created", "hour"]).agg(
        reg=("regular_volume", "sum"),
        vol=("volumetric_volume", "sum"),
    ).reset_index()

    rows = []
    for (ln, lt, dt), g in grp.groupby(["layout_name", "Layout Type", "Date of created"]):
        reg_h = {}; vol_h = {}
        for _, r in g.iterrows():
            h = int(r["hour"])
            reg_h[h] = float(r["reg"])
            vol_h[h] = float(r["vol"])
        rows.append({"layout_name": ln, "Layout Type": lt, "date": dt,
                      "reg": reg_h, "vol": vol_h})
    return rows


def _header_row(ws, left_title, right_title, left_fill, right_fill):
    """Write standard header."""
    ws.cell(1, COL_A, "Layout").font = HEADER_FONT
    ws.cell(1, COL_B, "Layout Type").font = HEADER_FONT
    ws.cell(1, COL_C, "Date").font = HEADER_FONT
    for h in range(24):
        c = ws.cell(1, LEFT_START + h, f"{left_title} {HOUR_LABELS[h]}")
        c.font = HEADER_FONT; c.fill = left_fill
    c = ws.cell(1, LEFT_TOTAL, f"{left_title} Total")
    c.font = HEADER_FONT; c.fill = left_fill
    ws.cell(1, SPACER, "").fill = SPACER_FILL
    for h in range(24):
        c = ws.cell(1, RIGHT_START + h, f"{right_title} {HOUR_LABELS[h]}")
        c.font = HEADER_FONT; c.fill = right_fill
    c = ws.cell(1, RIGHT_TOTAL, f"{right_title} Total")
    c.font = HEADER_FONT; c.fill = right_fill


def _idx_cells(ws, r, row):
    ws.cell(r, COL_A, row["layout_name"])
    ws.cell(r, COL_B, row["Layout Type"])
    ws.cell(r, COL_C, row["date"])


def _sum_formula(start_col, end_col, r):
    return f"=SUM({get_column_letter(start_col)}{r}:{get_column_letter(end_col)}{r})"


# ═══════════════════════════════════════════════════════════════
# Sheet writers
# ═══════════════════════════════════════════════════════════════

def _write_raw_data(wb, row_data):
    """Raw Data: Regular volume (left) | Volumetric volume (right). Hardcoded values."""
    ws = wb.create_sheet(title="Raw Data")
    _header_row(ws, "Reg Vol", "Vol Vol", BLUE_FILL, ORANGE_FILL)

    for i, row in enumerate(row_data):
        r = i + 2
        _idx_cells(ws, r, row)
        for h in range(24):
            ws.cell(r, LEFT_START + h, row["reg"].get(h, 0)).number_format = NUM_VOL
            ws.cell(r, RIGHT_START + h, row["vol"].get(h, 0)).number_format = NUM_VOL
        ws.cell(r, LEFT_TOTAL, _sum_formula(LEFT_START, LEFT_END, r)).number_format = NUM_VOL
        ws.cell(r, SPACER, "").fill = SPACER_FILL
        ws.cell(r, RIGHT_TOTAL, _sum_formula(RIGHT_START, RIGHT_END, r)).number_format = NUM_VOL

    ws.freeze_panes = "D2"


def _write_activity_sheet(wb, sheet_name, activity, row_data,
                           load_factors, effort, offsets, is_vol, xd_mult=None):
    """Activity sheet:
    LEFT  = time-shifted activity volume = FORMULA → 'Raw Data' × load_factor × offset_fraction
    RIGHT = MH = FORMULA → left_cell × effort (× XD multiplier if applicable)
    """
    ws = wb.create_sheet(title=sheet_name)
    mh_label = f"MH (×{xd_mult})" if xd_mult else "MH"
    _header_row(ws, "Act Vol", mh_label, BLUE_FILL, GREEN_FILL)

    # Which Raw Data block to reference: regular (LEFT) or volumetric (RIGHT)
    raw_block_start = RIGHT_START if is_vol else LEFT_START

    for i, row in enumerate(row_data):
        r = i + 2
        lt = row["Layout Type"]
        _idx_cells(ws, r, row)

        lf = 1.0 if is_vol else load_factors.get(lt, {}).get(activity, 0)

        # ── LEFT: Activity volume at target hour (time-shifted) ──
        # Formula: SUM over offsets of ('Raw Data'!source_hour_col × load_factor × fraction)
        for target_h in range(24):
            if lf == 0:
                ws.cell(r, LEFT_START + target_h, 0).number_format = NUM_VOL
            else:
                parts = []
                for (delta, frac) in offsets:
                    source_h = (target_h - delta) % 24
                    raw_col = get_column_letter(raw_block_start + source_h)
                    composite_lf = round(lf * frac, 10)
                    if composite_lf == 1.0:
                        parts.append(f"'Raw Data'!{raw_col}{r}")
                    elif composite_lf > 0:
                        parts.append(f"'Raw Data'!{raw_col}{r}*{composite_lf}")
                formula = "=" + "+".join(parts) if parts else "=0"
                ws.cell(r, LEFT_START + target_h, formula).number_format = NUM_VOL

        ws.cell(r, LEFT_TOTAL, _sum_formula(LEFT_START, LEFT_END, r)).number_format = NUM_VOL
        ws.cell(r, SPACER, "").fill = SPACER_FILL

        # ── RIGHT: MH = left_cell × effort ──
        for h in range(24):
            vol_col = get_column_letter(LEFT_START + h)
            if effort == 0 or lf == 0:
                ws.cell(r, RIGHT_START + h, 0).number_format = NUM_MH
            else:
                ws.cell(r, RIGHT_START + h, f"={vol_col}{r}*{round(effort, 10)}").number_format = NUM_MH

        ws.cell(r, RIGHT_TOTAL, _sum_formula(RIGHT_START, RIGHT_END, r)).number_format = NUM_MH

    ws.freeze_panes = "D2"


def _write_summary(wb, row_data, act_sheets, is_xd=False):
    """Summary: Raw volume (left, from Raw Data) | Total MH (right, sum of all activity MH)."""
    ws = wb.create_sheet(title="Summary")
    mh_title = "Total MH (XD Adjusted)" if is_xd else "Total MH"
    _header_row(ws, "Raw Vol", mh_title, BLUE_FILL, GREEN_FILL)

    sheet_names = list(act_sheets.keys())

    for i, row in enumerate(row_data):
        r = i + 2
        _idx_cells(ws, r, row)

        # LEFT: reference Raw Data regular volume
        for h in range(24):
            raw_col = get_column_letter(LEFT_START + h)
            ws.cell(r, LEFT_START + h, f"='Raw Data'!{raw_col}{r}").number_format = NUM_VOL

        ws.cell(r, LEFT_TOTAL, _sum_formula(LEFT_START, LEFT_END, r)).number_format = NUM_VOL
        ws.cell(r, SPACER, "").fill = SPACER_FILL

        # RIGHT: SUM of MH from all activity sheets (their RIGHT block)
        for h in range(24):
            mh_col = get_column_letter(RIGHT_START + h)
            parts = [f"'{sn.replace(chr(39), chr(39)+chr(39))}'!{mh_col}{r}"
                     for sn in sheet_names]
            formula = "=" + "+".join(parts) if parts else "=0"
            ws.cell(r, RIGHT_START + h, formula).number_format = NUM_MH

        ws.cell(r, RIGHT_TOTAL, _sum_formula(RIGHT_START, RIGHT_END, r)).number_format = NUM_MH

    ws.freeze_panes = "D2"


def _write_staffing(wb, staff, xd_staff=None):
    """Staffing summary with effective heads (perm + avg flex)."""
    ws = wb.create_sheet(title="Staffing")
    if staff.empty:
        ws["A1"] = "No staffing data"
        return

    cols = ["layout_name", "Layout Type", "shift", "peak_mh",
            "perm_heads", "flex_heads", "total_heads"]
    extra = [c for c in ["avg_daily_flex", "effective_heads", "avg_daily_vol", "productivity"]
             if c in staff.columns]
    cols += extra

    hmap = {
        "layout_name": "Layout", "Layout Type": "Layout Type", "shift": "Shift",
        "peak_mh": "Peak MH", "perm_heads": "Permanent", "flex_heads": "Peak Flex",
        "total_heads": "Peak Total", "avg_daily_flex": "Avg Daily Flex",
        "effective_heads": "Effective Heads", "avg_daily_vol": "Avg Daily Vol",
        "productivity": "Productivity",
    }
    for j, col in enumerate(cols, 1):
        c = ws.cell(1, j, hmap.get(col, col))
        c.font = HEADER_FONT; c.fill = BLUE_FILL

    out = staff[cols].sort_values(["layout_name", "shift"])
    for i, (_, row) in enumerate(out.iterrows(), 2):
        for j, col in enumerate(cols, 1):
            ws.cell(i, j, row[col])

    r = len(out) + 2
    ws.cell(r, 1, "TOTAL").font = HEADER_FONT
    sum_cols = ["peak_mh", "perm_heads", "flex_heads", "total_heads",
                "avg_daily_flex", "effective_heads", "avg_daily_vol"]
    for j, col in enumerate(cols, 1):
        if col in sum_cols:
            cl = get_column_letter(j)
            ws.cell(r, j, f"=SUM({cl}2:{cl}{r-1})")
            ws.cell(r, j).font = HEADER_FONT

    # XD Combined Staffing section (if applicable)
    if xd_staff is not None and not xd_staff.empty:
        r += 2
        ws.cell(r, 1, "XD COMBINED STAFFING (Dock + Processing)").font = HEADER_FONT
        ws.cell(r, 1).fill = ORANGE_FILL
        r += 1
        xd_cols = ["shift", "peak_mh", "perm_heads", "flex_heads", "total_heads"]
        xd_hmap = {"shift": "Shift", "peak_mh": "Combined Peak MH",
                    "perm_heads": "Permanent", "flex_heads": "Flex", "total_heads": "Total"}
        for j, col in enumerate(xd_cols, 1):
            c = ws.cell(r, j, xd_hmap.get(col, col))
            c.font = HEADER_FONT; c.fill = ORANGE_FILL
        r += 1
        xd_out = xd_staff[xd_cols].sort_values("shift")
        for _, row in xd_out.iterrows():
            for j, col in enumerate(xd_cols, 1):
                ws.cell(r, j, row[col])
            r += 1
        # Total
        ws.cell(r, 1, "XD TOTAL").font = HEADER_FONT
        for j, col in enumerate(xd_cols, 1):
            if col in ["peak_mh", "perm_heads", "flex_heads", "total_heads"]:
                cl = get_column_letter(j)
                ws.cell(r, j, f"=SUM({cl}{r-len(xd_out)}:{cl}{r-1})")
                ws.cell(r, j).font = HEADER_FONT

    ws.freeze_panes = "A2"
